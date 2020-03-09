import os
import sys
import csv
import struct
import re
import operator
import openpyxl


def main():
    inFile = os.path.abspath(sys.argv[1])
    basefile = os.path.basename(inFile)
    outputdir = os.path.dirname(inFile)
    
    staticdata = os.path.join(outputdir, "StaticData.dat")
    
    ## Dictionary for storing records for each category
    recordsDict = {"overnight" : {"DOM" : [], "FGN" : []},
                   6:{"DOM" : [], "MEX" : [], "CAN" : [], "OTHER" : []},
                   7:{"DOM" : [], "MEX" : [], "CAN" : [], "OTHER" : []}}
                   
    line_len = 1200  # Correct length of line to parse    
    
    with open(inFile, 'rb') as o:
        with open(staticdata, 'wb') as s:
        
            jobnumber = provideJobNumber()
            static_hdr = makeHeader()   # Static File header
            parse = makeLineParser()
            
            print "Writing Static Data and sorting foreign and domestic records..."
            writeStatic = csv.writer(s, quoting=csv.QUOTE_ALL)
            writeStatic.writerow(static_hdr)
    
            for seq, line in enumerate(o, start=1):
                
                ## Create static data line and write to file
                staticLine = createStaticLine(parse, line, line_len, seq)
                writeStatic.writerow(staticLine)
                
                ##  Create address line and Segregate records into respective categories
                addrLine = createAddrLine(staticLine, static_hdr)
                segregateRecords(addrLine, staticLine, static_hdr, recordsDict)
                
                if seq % 1000 == 0:
                    print "{} records sorted".format(seq)

            ## Output Address files
            # try:
                # print "\r\nWriting Overnight Records...." 
                # sortOvernightRecordsByName(recordsDict)
                # OvernightCSV = os.path.join(outputdir, "Overnight.csv")
                # writeAddressFile(OvernightCSV, recordsDict["overnight"])
                # createOverNightLabelsFile(recordsDict, outputdir, jobnumber)
            # except:
                # print "No Overnight records to process."

            print "\r\nWriting Overnight Records...." 
            sortOvernightRecordsByName(recordsDict)
            OvernightCSV = os.path.join(outputdir, "Overnight.csv")
            writeAddressFile(OvernightCSV, recordsDict["overnight"])
            createOverNightLabelsFile(recordsDict, outputdir, jobnumber)


            try:
                print "\r\nWriting 6 sheet Records..." 
                Address_6_Sheets_CSV = os.path.join(outputdir, "Address_6_Sheets.csv")
                writeAddressFile(Address_6_Sheets_CSV, recordsDict[6])
            except:
                print "No 6 Sheets records to process."
            
            try:
                print "\r\nWriting 7 sheet Records..." 
                Address_7_Sheets_CSV = os.path.join(outputdir, "Address_7_Sheets.csv")
                writeAddressFile(Address_7_Sheets_CSV, recordsDict[7])
            except:
                print "No 7 Sheets records to process."
            
            print "\r\nCreating Counts Report..." 
            # Get counts. Write to file and print to screen.
            countsreport = os.path.join(outputdir, "Counts.txt")
            createCountsReport(countsreport, recordsDict)
            
            print "\r\nCreating Overnight and Mail Excel File..." 
            createRecordsExcel(recordsDict, static_hdr, outputdir, jobnumber)
            


def provideJobNumber():
    """ Get 5 digit Mailshop job ticker number from input """
    jobnumber = raw_input("1. Provide DS job number:  ")
    while re.match(r'^\d{5}$', jobnumber) == None:
        jobnumber = raw_input("Job number is not valid !!! -->  ")
        
    return jobnumber 


def makeLineParser():
    """ Parser to split line into fields by field lengths. """

    formatStr = "5s 10s 40s 4s 15s 10s 15s 40s 40s 40s 40s 40s 40s 15s 15s 15s 40s 10s 120s 220s 80s 190s 1s 1s 10s 144s"
    fieldstruct = struct.Struct(formatStr)
    parse = fieldstruct.unpack_from
    return parse


def makeHeader():
    return ["Company Number","Account Number","Company Name",
            "Form Code","Shares Lost","Item Number","Phone Number",
            "NameAddress1","NameAddress2","NameAddress3",
            "NameAddress4","NameAddress5","NameAddress6",
            "Total Fee","Price Per Share","Shares to Sell",
            "Country","Mail Date","Shareholder Name","Certificates",
            "Company Name ","Shareholder Address","Main Document Indicator",
            "Additional Document Indicator","Sequence number","Filler"]


def createStaticLine(parse, line, line_len, seq):
    """ Convert byte text to unicode chars. Replace non-ASCII,
    the "replacement", "non-breaking space" and "Broken Bar" 
    chars. Convert chars back into bytes. """
    
    char_text = line.decode("utf-8", errors='replace').replace(u'\ufffd', " ")
    replace_nbspace = char_text.replace(u'\u00A6', " ")
    replace_bknbar = replace_nbspace.replace(u'\u00A0', " ")
    latin_line = replace_bknbar.encode("latin-1")
    latin_text = latin_line.decode("ascii", errors='replace').replace(u'\ufffd', " ")
    ascii_line = latin_text.encode("ascii")
    
    if len(ascii_line) != line_len:
        print "Incorrect length: Line {} ({} chars) should be {}".format(seq, len(ascii_line), line_len)
        print line
    
    staticLine = [" ".join(x.split()) for x in parse(ascii_line)]
    
    return staticLine    
    
 
def createAddrLine(staticLine, static_hdr):
    """ Extract address fields from parsed line. 
    Improve address processing in Mail Manager by 
    making sure all street address fields are found. 
    Check if APT, UNIT, FLOOR, etc is in another field. """
    
    addrline = staticLine[static_hdr.index("NameAddress1"):static_hdr.index("NameAddress6")+1]
    newAddr = []
    
    # If record s foreign, do not shift any address lines
    # All lines will be imported into BCC into the name line fields only
    # Prevents any potential reformatting by Mail Manager
    if staticLine[static_hdr.index("Country")] != "":
        newAddr = addrline + ([""] * 5)
    
    else:    
        addr = [f for f in addrline if f.upper() not in ["","NULL","."]]
        
        apt_pattern = re.compile(r'^((#|B(UI)?LD(IN)?G|S(UI)?TE|LOT|UNIT|FLOOR|R(OO)?M|AP(ARTMEN)?T|ATTN|\().+|(\d{1,4}\s?\w)|(\d{1,3}(ST|ND|RD|TH)?\s?FL(OO)?R?))$', flags=re.IGNORECASE)
    
        idx = -3 if apt_pattern.match(addr[-2]) and len(addr) > 3 != None else -2
        
        nameLines = addr[:idx]
        spaces = [""] * (len(addrline)-len(nameLines))
        streetAddr = addr[idx]
        alternateAddr = addr[-2] if apt_pattern.match(addr[-2]) and len(addr) > 3 != None else "" 
        cityStateZip = addr[-1]

        # 2 extra spaces at the end to pad the STATE and ZIP fields
        stateAndZipPlaceholders = ["", ""]

        newAddr = nameLines + spaces + [streetAddr, alternateAddr] + [cityStateZip] + stateAndZipPlaceholders

    compno_acctno_ltrno = [
    staticLine[static_hdr.index("Company Number")],
    staticLine[static_hdr.index("Account Number")],
    staticLine[static_hdr.index("Main Document Indicator")]
    ]

    addrLine = newAddr + compno_acctno_ltrno
    return addrLine


def segregateRecords(addrLine, staticLine, static_hdr, recordsDict):    
    """ Segregate domestic, Mexico, Canada and other foreign records."""

    bond_fee = staticLine[static_hdr.index("Total Fee")]
    wholeNumber = int(bond_fee.split(".")[0])

    if wholeNumber >= 10000:
        addrLine.append("Overnight")
        if staticLine[static_hdr.index("Country")] != "":
            recordsDict["overnight"]["FGN"].append((addrLine, staticLine))
        else:    
            recordsDict["overnight"]["DOM"].append((addrLine, staticLine))
    else:
        if staticLine[static_hdr.index("Additional Document Indicator")] in ["1","2"]:
            addrLine.append("7 Sheets")
            if staticLine[static_hdr.index("Country")] == "":
                recordsDict[7]["DOM"].append((addrLine, staticLine))
            elif staticLine[static_hdr.index("Country")] == "CANADA":
                recordsDict[7]["CAN"].append((addrLine, staticLine))
            elif staticLine[static_hdr.index("Country")] == "MEXICO":
                recordsDict[7]["MEX"].append((addrLine, staticLine))
            else:
                recordsDict[7]["OTHER"].append((addrLine, staticLine))
        else:
            addrLine.append("6 Sheets")
            if staticLine[static_hdr.index("Country")] == "":
                recordsDict[6]["DOM"].append((addrLine, staticLine))
            elif staticLine[static_hdr.index("Country")] == "CANADA":
                recordsDict[6]["CAN"].append((addrLine, staticLine))
            elif staticLine[static_hdr.index("Country")] == "MEXICO":
                recordsDict[6]["MEX"].append((addrLine, staticLine))
            else:
                recordsDict[6]["OTHER"].append((addrLine, staticLine))
                
                
def sortOvernightRecordsByName(recordsDict):
    """ Sort overnight records by the Shareholders name. """
    
    d = recordsDict["overnight"]["DOM"]
    recordsDict["overnight"]["DOM"] = sorted(d, key=lambda d_tup: d_tup[0][0])   

    f = recordsDict["overnight"]["FGN"]
    recordsDict["overnight"]["FGN"] = sorted(f, key=lambda f_tup: f_tup[0][0])       
   

def writeAddressFile(file, records_group):
    """ Format address data as straight First Class 
    and write to file. """

    domestic_records = records_group["DOM"]
    
    mexico = records_group.get("MEX") if records_group.get("MEX") != None else []
    canada = records_group.get("CAN") if records_group.get("CAN") != None else []
    other = records_group.get("OTHER") if records_group.get("OTHER") != None else []
    fgn = records_group.get("FGN") if records_group.get("FGN") != None else []
    
    foreign_records = mexico + canada + other + fgn
    
    addr_hdr = ["IM barcode Digits","OEL",
                "Sack and Pack Numbers","Presort Sequence",
                "Full Name","Name2","Name3",
                "Name4","Name5","Name6",
                "Delivery Address","Alternate 1 Address",
                "City","State","ZIP+4",
                "CompanyNo","AccountNo",
                "Letter", "PrintCategory"]

    with open(file, 'wb') as f:
        csvWriter = csv.writer(f, quoting=csv.QUOTE_ALL)
        csvWriter.writerow(addr_hdr)
        
        if len(foreign_records) > 0:
            addSeqNoAndWriteRecords(foreign_records, csvWriter)
        if len(domestic_records) > 0:
            addSeqNoAndWriteRecords(domestic_records, csvWriter)


def addSeqNoAndWriteRecords(records, csvWriter):
        
    for seq, tup in enumerate(records, start=1):
        addrLine = tup[0]
        
        seqNo = "{}".format(seq)
        
        imb_oel_sackpack_seq = ["","","", seqNo]
        newaddrline = imb_oel_sackpack_seq + addrLine
        csvWriter.writerow(newaddrline)
    

def createOverNightLabelsFile(recordsDict, outputdir, jobnumber):
    
    dom_hdr = ["FullName","DeliveryAddress","City","StateProvince",
              "PostalCode","Country","REFERENCE 1","REFERENCE 2"]
              
    fgn_hdr = ["FullName","Name2","Name3","Name4","Name5","Name6",
              "DeliveryAddress","City","StateProvince",
              "PostalCode","Country","REFERENCE 1","REFERENCE 2"]

    COUNTRY = ""
    REF_1 = "DEPT 152"
    REF_2 = "{} Aged Loss".format(jobnumber)
    
    # Process Domestic Records
    wb = openpyxl.Workbook()
    ws_dom = wb.create_sheet("Domestic", 0)
    ws_dom.append(dom_hdr)
    
    state_pattern = re.compile(r'(.+)(\s?\w\s?\w)(?=(\s\d{4,5}(\s|-)?(\d+)?))')
    
    for tup in recordsDict["overnight"]["DOM"]:
        addrLine = tup[0]
        domesticRow = formatDomesticRow(addrLine, state_pattern)
        ws_dom.append(domesticRow + [COUNTRY, REF_1, REF_2])

    # Process Foreign Records
    ws_fgn = wb.create_sheet("Foreign", 1) 
    ws_fgn.append(fgn_hdr)
    for tup in recordsDict["overnight"]["FGN"]:
        addrLine = tup[0]
        foreignRow = formatForeignRow(addrLine)
        ws_fgn.append(foreignRow + [COUNTRY, REF_1, REF_2])

    # Save Excel workbook
    overnight_label_file = os.path.join(outputdir, "OVERNIGHT_LABELS.xlsx")
    wb.save(overnight_label_file)


def formatDomesticRow(addrLine, state_pattern):

    ''' addrLine is  
        [Full Name, Name2, Name3, Name4, Name5, Name6, 
        Addr1, Addr2, City, State, Zip,
        CompNo, AcctNo, LtrNo, PrintCategory] '''

    addrLines = addrLine[:11]
    name = " ".join(addrLines[0:2])
    addr = " ".join(addrLines[6:8])
    cityStateZip = addrLines[8]
    
    state_abbrv_search = state_pattern.search(cityStateZip)
    city = ' '.join(state_abbrv_search.group(1).strip().upper().split()).strip(",") if state_abbrv_search != None else cityStateZip
    state = ' '.join(state_abbrv_search.group(2).strip().upper().split()) if state_abbrv_search != None else ""
    zip = ' '.join(state_abbrv_search.group(3).strip().upper().split()) if state_abbrv_search != None else ""
    
    return [name,addr,city,state,zip]
    

def formatForeignRow(addrLine):
    addrLines = addrLine[:6]
    return addrLines + ["","","",""]


def createCountsReport(countsreport, recordsDict):

    Overnight_Domestic = len(recordsDict["overnight"]["DOM"])
    Overnight_Foreign = len(recordsDict["overnight"]["FGN"])
    Total_Overnight = Overnight_Domestic + Overnight_Foreign
    
    Domestic_6pg = len(recordsDict[6]["DOM"])
    mexico_6pg  = len(recordsDict[6]["MEX"])
    canada_6pg  = len(recordsDict[6]["CAN"])
    other_6pg  = len(recordsDict[6]["OTHER"])
    Foreign_6pg  = mexico_6pg + canada_6pg + other_6pg
    Total_6pg    = Domestic_6pg + Foreign_6pg
    
    Domestic_7pg = len(recordsDict[7]["DOM"])
    mexico_7pg  = len(recordsDict[7]["MEX"])
    canada_7pg  = len(recordsDict[7]["CAN"])
    other_7pg  = len(recordsDict[7]["OTHER"])
    Foreign_7pg  = mexico_7pg + canada_7pg + other_7pg
    Total_7pg    = Domestic_7pg + Foreign_7pg
    
    recordCount = Total_Overnight + Total_6pg + Total_7pg 
    
    report = "\r\n".join([
        "Total Records: {}".format(recordCount),
        "",
        "Overnight Records (Bond Fee over 10K): {}".format(Total_Overnight),
        "  Domestic: {}     Foreign: {}".format(Overnight_Domestic, Overnight_Foreign),
        "",
        "6 Sheets Records (Bond Fee below 10K): {}".format(Total_6pg),
        "  Domestic: {}     Foreign: {}".format(Domestic_6pg, Foreign_6pg),
        "                   ---------------",
        "                   Mexico: {}    Canada: {}    Other: {}".format(mexico_6pg, canada_6pg, other_6pg),
        "",
        "7 Sheets Records (Bond Fee below 10K): {}".format(Total_7pg),
        "  Domestic: {}     Foreign: {}".format(Domestic_7pg, Foreign_7pg),     
        "                   ---------------",
        "                   Mexico: {}    Canada: {}    Other: {}".format(mexico_7pg, canada_7pg, other_7pg)
    ])
    
    with open(countsreport, 'wb') as c:
        c.write(report)
        print report
    
    
def createRecordsExcel(recordsDict, static_hdr, outputdir, jobnumber):
    
    wb = openpyxl.Workbook()
    
    # Process Overnight Records
    ws_ovr = wb.create_sheet("Overnight", 0)
    ws_ovr.append(static_hdr)
    
    try:
        # Combine records and resort by Company No and Account No
        overnight_records = recordsDict["overnight"]["DOM"] + recordsDict["overnight"]["FGN"]
        overnight_sorted = sorted(overnight_records, key=lambda tup: "{}{}".format(tup[1][0], tup[1][1]))
        for addr, static in overnight_sorted:
            ws_ovr.append(static)
    except:
        print "No Overnight records written to Excel"

    
    # Process Mail Records
    ws_mail = wb.create_sheet("Mailed", 1)
    ws_mail.append(static_hdr)

    try:
        # Combine records and resort by Company No and Account No
        records_6_sheets = recordsDict[6]["DOM"] + recordsDict[6]["MEX"] + recordsDict[6]["CAN"] + recordsDict[6]["OTHER"]  
        records_7_sheets = recordsDict[7]["DOM"] + recordsDict[7]["MEX"] + recordsDict[7]["CAN"] + recordsDict[7]["OTHER"]
        mail_records =  records_6_sheets + records_7_sheets
        mail_sorted = sorted(mail_records, key=lambda tup: "{}{}".format(tup[1][0], tup[1][1]))
        for addr, static in mail_sorted:
            ws_mail.append(static)
    except:
        print "No Mailed records written to Excel"
    
    # Save Excel workbook
    records_file = os.path.join(outputdir, "{} Aged Loss - Records.xlsx".format(jobnumber))
    wb.save(records_file)
            
if __name__ == "__main__":
    main()
